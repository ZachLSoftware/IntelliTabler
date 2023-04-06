from .models import *
import tempfile
import pandas as pd
from openpyxl import load_workbook
import xlsxwriter
import io
import numpy as np

def moduleTeacherWSTemplate(timetable, sheets):
    template_builders = {
        'Classes': moduleParentTemplateBuilder,
        'Teachers': teacherTemplateBuilder,
        'Schedule': scheduleTemplateBuilder,
        'Preferences': preferenceTemplateBuilder,
        'Assignments': assignmentTemplateBuilder
    }

    template= io.BytesIO()
    writer=pd.ExcelWriter(template, engine='xlsxwriter') 
    workbook = writer.book
    lockCell = workbook.add_format({'locked': True, 'num_format': ';;;'})
    for sheet in sheets:
        template_builders[sheet](timetable, workbook.add_worksheet(sheet), writer, lockCell)
    writer.close()
    return template.getvalue()

def moduleParentTemplateBuilder(timetable, class_sheet, writer, lock):
    classes=ModuleParent.objects.filter(timetable=timetable).order_by('name')
    classPd=pd.DataFrame(columns=['Class Name', 'Num Periods Taught', 'Num Class Groups', 'id'])
    if classes:
        for cl in classes:
            classPd.loc[len(classPd)] = [cl.name, cl.numPeriods, cl.numClasses, cl.id]
    classPd.to_excel(writer, sheet_name="Classes", index=False)
    class_sheet.set_column('D:D', None, lock, {'hidden': True})
    

def teacherTemplateBuilder(timetable, teacher_sheet, writer, lock):
    teachers=Teacher.objects.filter(department=timetable.tableYear.department).order_by('name')
    teacherPd=pd.DataFrame(columns=['Name', 'Load Per Week', 'Room', 'id'])
    
    if teachers:
        for teacher in teachers:
            teacherPd.loc[len(teacherPd)]=[teacher.name, teacher.load, teacher.roomNum, teacher.id]
    teacherPd.to_excel(writer, sheet_name="Teachers", index=False)
    teacher_sheet.set_column('D:D', None, lock, {'hidden': True})

def preferenceTemplateBuilder(timetable, preference_sheet, writer, lock):
    moduleList=list(set(Module.objects.values_list('name', flat=True).filter(group__parent__timetable=timetable)))
    moduleList.sort()
    teachers=list(Teacher.objects.values_list('name', flat=True).filter(department=timetable.tableYear.department).order_by('name'))
    preference=['NONE','MEDIUM','HIGH','REQUIRED']
    preferenceDF=pd.DataFrame(columns=['Class', 'Teacher', 'Priority'])
    preferenceDF.to_excel(writer,sheet_name="Preferences", index=False)

    moduleValidation = {
        'validate': 'list',
        'source': moduleList,
        'input_title': 'Choose a Module',
        'input_message': 'Choose from the dropdown list.',
        'show_input': True,
        'show_error': True,
        'error_title': 'Invalid Input',
        'error_message': 'Please select a valid value from the dropdown list.',
        'error_type': 'information',
        'ignore_blank': False
    }
    teacherValidation = {
        'validate': 'list',
        'source': teachers,
        'input_title': 'Select a teacher',
        'input_message': 'Choose from the dropdown list.',
        'show_input': True,
        'show_error': True,
        'error_title': 'Invalid Input',
        'error_message': 'Please select a valid value from the dropdown list.',
        'error_type': 'information',
        'ignore_blank': False
    }
    preferenceValidation={
        'validate': 'list',
        'source': preference,
        'input_title': 'Select a Priority',
        'input_message': 'Choose from the dropdown list.',
        'show_input': True,
        'show_error': True,
        'error_title': 'Invalid Input',
        'error_message': 'Please select a valid value from the dropdown list.',
        'error_type': 'information',
        'ignore_blank': False
    }
    preference_sheet.data_validation(1,0, 200,0, moduleValidation)
    preference_sheet.data_validation(1,1, 200,1, teacherValidation)
    preference_sheet.data_validation(1,2, 200,2, preferenceValidation)

def assignmentTemplateBuilder(timetable, assignment_sheet, writer, lock):
    moduleList=list(set(Module.objects.values_list('name', flat=True).filter(group__parent__timetable=timetable)))
    moduleList.sort()
    teachers=list(Teacher.objects.values_list('name', flat=True).filter(department=timetable.tableYear.department).order_by('name'))
    assignmentDF=pd.DataFrame(columns=['Class', 'Teacher'])
    assignmentDF.to_excel(writer,sheet_name="Assignments", index=False)

    moduleValidation = {
        'validate': 'list',
        'source': moduleList,
        'input_title': 'Choose a Module',
        'input_message': 'Choose from the dropdown list.',
        'show_input': True,
        'show_error': True,
        'error_title': 'Invalid Input',
        'error_message': 'Please select a valid value from the dropdown list.',
        'error_type': 'information',
        'ignore_blank': False
    }
    teacherValidation = {
        'validate': 'list',
        'source': teachers,
        'input_title': 'Select a teacher',
        'input_message': 'Choose from the dropdown list.',
        'show_input': True,
        'show_error': True,
        'error_title': 'Invalid Input',
        'error_message': 'Please select a valid value from the dropdown list.',
        'error_type': 'information',
        'ignore_blank': False
    }
    
    assignment_sheet.data_validation(1,0, 200,0, moduleValidation)
    assignment_sheet.data_validation(1,1, 200,1, teacherValidation)

def scheduleTemplateBuilder(timetable, timeslots_sheet, writer, lock):
    periods=list(Period.objects.filter(department=timetable.tableYear.department, week=1).order_by('dayNum').values_list('name', flat=True))
    w=timetable.tableYear.department.format.numWeeks
    groups=ModuleGroup.objects.filter(parent__timetable=timetable).order_by('name')
    groupPd=pd.DataFrame(columns=["Name","Period", "Week", "id"])
    if groups:
        for group in groups:
            groupPd.loc[len(groupPd)] = [group.name, group.period.name if group.period else '', group.period.week if group.period else '', group.id]
    groupPd.to_excel(writer, sheet_name="Schedule", index=False)
    
    periodValidation = {
        'validate': 'list',
        'source': periods,
        'input_title': 'Enter a value:',
        'input_message': 'Choose from the dropdown list.',
        'show_input': True,
        'show_error': True,
        'error_title': 'Invalid Input',
        'error_message': 'Please select a valid value from the dropdown list.',
        'error_type': 'information',
        'ignore_blank': True
    }
    weekValidation = {
        'validate': 'integer',
        'criteria': 'between',
        'minimum': 1,
        'maximum': w,
        'input_title': 'Enter a value:',
        'input_message': 'between 1 and ' + str(w),
        'show_input': True,
        'show_error': True,
        'error_title': 'Invalid Input',
        'error_message': 'Please select a valid value from the dropdown list.',
        'error_type': 'information',
        'ignore_blank': True
    }
    timeslots_sheet.set_column('D:D', None, lock, {'hidden': True})
    timeslots_sheet.data_validation(1,1, len(groupPd),1, periodValidation)
    timeslots_sheet.data_validation(1,2, len(groupPd),2, weekValidation)


def readInExcel(timetable, ws):
    workbook = load_workbook(ws)
    sheets=workbook.sheetnames
    template_readers = {
        'Classes': moduleParentTemplateReader,
        'Teachers': teacherTemplateReader,
        'Schedule': scheduleTemplateReader,
        'Preferences': preferencesTemplateReader,
        'Assignments': assignmentsTemplateReader
    }
    results={}
    for sheet in sheets:
        results[sheet]=template_readers[sheet](timetable, pd.read_excel(ws, sheet_name=sheet))
        
    return results

def moduleParentTemplateReader(timetable, classesDF):
    classDF=classDF.replace({0:np.NaN, '':np.NaN, ' ':np.NaN})
    for i, r in classDF.iterrows():
        if not pd.isnull(r['Class Name']) and not pd.isnull(r['Num Periods Taught']) and not pd.isnull(r['Num Class Groups']):
            if not pd.isnull(r['id']):
                try:
                    cl=ModuleParent.objects.get(id=r['id'])
                except:
                    try:
                        cl=ModuleParent.objects.get(name=r['Class Name'], timetable=timetable)
                    except:
                        continue
            if not cl:
                cl=ModuleParent()
                cl.timetable=timetable
                cl.department=timetable.tableYear.department
                cl.user=timetable.user
            cl.name=r['Class Name']
            cl.numPeriods=r['Num Periods Taught']
            cl.numClasses=r['Num Class Groups']
            cl.save()
    return True

def teacherTemplateReader(timetable, teachersDF):

    teachersDF=teachersDF.replace({0:np.NaN, '':np.NaN, ' ':np.NaN})
    for i, r in teachersDF.iterrows():
        if not pd.isnull(r['Name']) and not pd.isnull(r['Load Per Week']):
            if not pd.isnull(r['id']):
                try:
                    teacher=Teacher.objects.get(id=r['id'])
                except:
                    try:
                        teacher=Teacher.objects.get(name=r['Name'], department=timetable.tableYear.department)
                    except:
                        teacher=None
            if not teacher:
                teacher=Teacher()
                teacher.department=timetable.tableYear.department
                teacher.user=timetable.user
            teacher.name=r['Name']
            teacher.load=r['Load Per Week']
            teacher.roomNum=r['Room']
            teacher.save()
    return True

def scheduleTemplateReader(timetable, timetableDF):
    timetableDF=timetableDF.replace({0:np.NaN, '':np.NaN, ' ':np.NaN})
    for i, r in timetableDF.iterrows():
        if not pd.isnull(r['Name']) and not pd.isnull(r['Period']) and not pd.isnull(r['Week']):
            try:
                mod=ModuleGroup.objects.get(id=r['id'])
            except:
                try:
                    mod=ModuleGroup.objects.get(name=r['Name'], parent__timetable=timetable)
                except:
                    continue
            
            mod.period=Period.objects.get(department=timetable.tableYear.department, name=r['Period'], week=r['Week'])
            mod.save()
    return True

def preferencesTemplateReader(timetable, preferencesDF):
    priorities={'NONE':0,'MEDIUM':1,'HIGH':2,'REQUIRED':3}
    preferencesDF=preferencesDF.replace({0:np.NaN, '':np.NaN, ' ':np.NaN})
    for i, r in preferencesDF.iterrows():
        if not pd.isnull(r['Class']) and not pd.isnull(r['Teacher']) and not pd.isnull(r['Priority']):
            try:
                mods=Module.objects.filter(name=r['Class'], group__parent__timetable=timetable)
                teacher=Teacher.objects.get(name=r['Teacher'], department=timetable.tableYear.department)
            except:
                continue
            try:
                for mod in mods:
                    Preference.objects.create(teacher=teacher, module=mod, priority=priorities[r['Priority']], timetable=timetable)
            except:
                continue       
    return True

def assignmentsTemplateReader(timetable, assignmentDF):
    assignmentDF=assignmentDF.replace({0:np.NaN, '':np.NaN, ' ':np.NaN})
    for i, r in assignmentDF.iterrows():
        if not pd.isnull(r['Class']) and not pd.isnull(r['Teacher']):
            try:
                teacher=Teacher.objects.get(name=r['Teacher'], department=timetable.tableYear.department)
                Module.objects.filter(name=r['Class'], group__parent__timetable=timetable).update(teacher=teacher)                
            except:
                continue
    return True

def combingTemplate(timetable):
    import os
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    file_path = os.path.join(BASE_DIR, 'static', 'excel_templates', 'combingTemplate.xlsx')

    workbook = load_workbook(filename=file_path)

    combing_sheet = workbook['Combing']
    teachers_sheet = workbook['Teachers']
    classes_sheet = workbook['Classes']
    format_sheet = workbook['Format']
    assignments_sheet = workbook['Assignments']

    format_sheet.append([timetable.tableYear.department.format.numPeriods, timetable.tableYear.department.format.numWeeks])

    teachers=Teacher.objects.filter(department=timetable.tableYear.department)
    classes=ModuleParent.objects.filter(timetable=timetable)
    modules=Module.objects.filter(group__parent__timetable=timetable,group__period__isnull=False, teacher__isnull=False)

    for teacher in teachers:
        teachers_sheet.append([teacher.name, teacher.load])
    
    for cl in classes:
        classes_sheet.append([cl.name, cl.numPeriods, cl.numClasses])

    for mod in modules:
        assignments_sheet.append([mod.teacher.name, mod.group.period.dayNum, mod.name])

    file = io.BytesIO()
    workbook.save(file)
    file.seek(0)

    return file.getvalue()

def readFromCombing(timetable, ws):
    combing_sheet = pd.read_excel(ws,sheet_name='Combing')
    teachers_sheet = pd.read_excel(ws,sheet_name='Teachers')
    classes_sheet = pd.read_excel(ws, sheet_name='Classes')
    format_sheet = pd.read_excel(ws, sheet_name='Format')

    teacherNullCheck=teachers_sheet.replace({"":np.NaN, 0:np.NaN}).copy()
    classesNullCheck=classes_sheet.replace({"":np.NaN, 0:np.NaN}).copy()

    if teacherNullCheck.isnull().sum().sum()>0 or classesNullCheck.isnull().sum().sum()>0:
        return False
    
    for i, r in classes_sheet.iterrows():
        try:
            cl=ModuleParent.objects.get(name=r['Class'], timetable=timetable)
        except:
            cl=None
        if not cl:
            cl=ModuleParent()
            cl.timetable=timetable
            cl.department=timetable.tableYear.department
            cl.user=timetable.user
        cl.name=r['Class']
        cl.numPeriods=r['Groups']
        cl.numClasses=r['Periods']
        cl.save()

    for i, r in teachers_sheet.iterrows():
        try:
            teacher=Teacher.objects.get(name=r['Teacher'], department=timetable.tableYear.department)
        except:
            teacher=None
        if not teacher:
            teacher=Teacher()
            teacher.department=timetable.tableYear.department
            teacher.user=timetable.user
        teacher.name=r['Teacher']
        teacher.load=r['Load Per Week']
        teacher.save()
    periodLen=timetable.tableYear.department.format.numPeriods*5*timetable.tableYear.department.format.numWeeks
    columns=periodLen+2
    readCombing=combing_sheet.replace(0,np.NaN).copy()
    readCombing=readCombing.iloc[:,0:columns]
    assignments={i:[] for i in range(1,periodLen+1)}
    sessions={}
    print(assignments)
    count=0
    print(readCombing.head())
    try:
        for index, row in readCombing.iterrows():
            if row['Teacher']=='Class' and row['Load']=="Periods":
                break
            else:
                try:
                    teacher=Teacher.objects.get(name=row['Teacher'], department=timetable.tableYear.department)
                except:
                    continue
            for column_name, value in row.items():
                if column_name not in ['Teacher', 'Load'] and pd.notna(value):
                    assignments[column_name].append((value,teacher))
                    sessions.setdefault(value, 1)

        
        for k,v in assignments.items():
            for tup in v:
                Module.objects.filter(group__parent__timetable=timetable, group__session=sessions[tup[0]], name=tup[0]).update(teacher=tup[1])
                sessions[tup[0]]+=1
    except Exception as e:
        print(e)
    return True

def exportCalendar(timetable, teacher=None):
    if teacher:
        classes=Module.objects.filter(group__parent__timetable=timetable, teacher=teacher, group__period__isnull=False, group__period__week=1).order_by('group', 'group__session', 'lesson')
    else:
        classes=ModuleGroup.objects.filter(parent__timetable=timetable, period__isnull=False, period__week=1).order_by('session')

    numPeriods=timetable.tableYear.department.format.numPeriods
    
    calendar = io.BytesIO()
    workbook=xlsxwriter.Workbook(calendar)
    

    worksheet=workbook.add_worksheet("Week 1")
    x={'Mon':1, 'Tues':2, 'Wed':3, 'Thurs':4, 'Fri':5}
    #worksheet.write_row(0, 1, header)

    write_dict={}
    wrap = workbook.add_format({'text_wrap': True})
    if teacher:
        for cl in classes:
            coord=cl.group.period.name.split('-')
            write_dict.setdefault((int(coord[1]),x[coord[0]]),'')
            write_dict[(int(coord[1]),x[coord[0]])]+=(cl.name+'\n')
    else:
        for cl in classes:
            coord=cl.period.name.split('-')
            write_dict.setdefault((int(coord[1]),x[coord[0]]),'')
            write_dict[(int(coord[1]),x[coord[0]])]+=(cl.name+'\n')
    for k,v in write_dict.items():
        worksheet.write(k[0],k[1], v, wrap)
    for i in range(2,numPeriods+2):
        row='A'+str(i)
        worksheet.write_formula(row, str(i-1))

    worksheet.set_column('B:F', 30)
    worksheet.add_table('A1:F6', {'columns': [{'header': 'Period'},
                                          {'header': 'Mon'},
                                          {'header': 'Tues'},
                                          {'header': 'Wed'},
                                          {'header': 'Thurs'},
                                          {'header': 'Fri'},
                                          ]})
    
    workbook.close()
    return calendar.getvalue()
    