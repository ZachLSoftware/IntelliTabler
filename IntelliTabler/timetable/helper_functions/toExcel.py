from ..models import *
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import xlsxwriter
import io
import numpy as np

"""
Runs through each template to be built.
"""
def templateBuilderMain(timetable, sheets):
    #Set template builders dict
    template_builders = {
        'Classes': moduleParentTemplateBuilder,
        'Teachers': teacherTemplateBuilder,
        'Schedule': scheduleTemplateBuilder,
        'Preferences': preferenceTemplateBuilder,
        'Assignments': assignmentTemplateBuilder
    }

    #create IO object to save file
    template= io.BytesIO()
    writer=pd.ExcelWriter(template, engine='xlsxwriter') #Create writer with pandas
    workbook = writer.book #Get workbook
    lockCell = workbook.add_format({'locked': True, 'num_format': ';;;'}) #Set locking cell format
    
    #For each sheet selected, run the template builder
    for sheet in sheets:
        template_builders[sheet](timetable, workbook.add_worksheet(sheet), writer, lockCell)

    #close the writer and return the IO object
    writer.close()
    return template.getvalue()

"""
Builds class template
"""
def moduleParentTemplateBuilder(timetable, class_sheet, writer, lock):
    classes=ModuleParent.objects.filter(timetable=timetable).order_by('name')
    classPd=pd.DataFrame(columns=['Class Name', 'Num Periods Taught', 'Num Class Groups', 'id'])
    if classes:
        for cl in classes:
            classPd.loc[len(classPd)] = [cl.name, cl.numPeriods, cl.numClasses, cl.id]
    classPd.to_excel(writer, sheet_name="Classes", index=False)
    class_sheet.set_column('D:D', None, lock, {'hidden': True})
    
"""
Builds teacher template
"""
def teacherTemplateBuilder(timetable, teacher_sheet, writer, lock):
    teachers=Teacher.objects.filter(department=timetable.tableYear.department).order_by('name')
    teacherPd=pd.DataFrame(columns=['Name', 'Load Per Week', 'Room', 'id'])
    
    if teachers:
        for teacher in teachers:
            teacherPd.loc[len(teacherPd)]=[teacher.name, teacher.load, teacher.roomNum, teacher.id]
    teacherPd.to_excel(writer, sheet_name="Teachers", index=False)
    teacher_sheet.set_column('D:D', None, lock, {'hidden': True})

"""
Builds preference template
"""
def preferenceTemplateBuilder(timetable, preference_sheet, writer, lock):
    moduleList=list(set(Module.objects.values_list('name', flat=True).filter(group__parent__timetable=timetable)))
    moduleList.sort()
    teachers=list(Teacher.objects.values_list('name', flat=True).filter(department=timetable.tableYear.department).order_by('name'))
    preference=['NONE','MEDIUM','HIGH','REQUIRED']
    preferenceDF=pd.DataFrame(columns=['Class', 'Teacher', 'Priority'])
    preferenceDF.to_excel(writer,sheet_name="Preferences", index=False)

    #Create dropdowns inside excel
    moduleValidation = {
        'validate': 'list',
        'source': moduleList,
        'input_title': 'Choose a Module',
        'input_message': 'Choose from the dropdown list.',
        'show_input': True,
        'show_error': True,
        'error_title': 'Invalid Input',
        'error_message': 'Please select a valid value from the dropdown list.',
        'error_type': 'information',
        'ignore_blank': False
    }
    teacherValidation = {
        'validate': 'list',
        'source': teachers,
        'input_title': 'Select a teacher',
        'input_message': 'Choose from the dropdown list.',
        'show_input': True,
        'show_error': True,
        'error_title': 'Invalid Input',
        'error_message': 'Please select a valid value from the dropdown list.',
        'error_type': 'information',
        'ignore_blank': False
    }
    preferenceValidation={
        'validate': 'list',
        'source': preference,
        'input_title': 'Select a Priority',
        'input_message': 'Choose from the dropdown list.',
        'show_input': True,
        'show_error': True,
        'error_title': 'Invalid Input',
        'error_message': 'Please select a valid value from the dropdown list.',
        'error_type': 'information',
        'ignore_blank': False
    }
    preference_sheet.data_validation(1,0, 200,0, moduleValidation)
    preference_sheet.data_validation(1,1, 200,1, teacherValidation)
    preference_sheet.data_validation(1,2, 200,2, preferenceValidation)

"""
Builds assignment template
"""
def assignmentTemplateBuilder(timetable, assignment_sheet, writer, lock):
    moduleList=list(set(Module.objects.values_list('name', flat=True).filter(group__parent__timetable=timetable)))
    moduleList.sort()
    teachers=list(Teacher.objects.values_list('name', flat=True).filter(department=timetable.tableYear.department).order_by('name'))
    assignmentDF=pd.DataFrame(columns=['Class', 'Teacher'])
    assignmentDF.to_excel(writer,sheet_name="Assignments", index=False)

    #create dropdowns inside excel
    moduleValidation = {
        'validate': 'list',
        'source': moduleList,
        'input_title': 'Choose a Module',
        'input_message': 'Choose from the dropdown list.',
        'show_input': True,
        'show_error': True,
        'error_title': 'Invalid Input',
        'error_message': 'Please select a valid value from the dropdown list.',
        'error_type': 'information',
        'ignore_blank': False
    }
    teacherValidation = {
        'validate': 'list',
        'source': teachers,
        'input_title': 'Select a teacher',
        'input_message': 'Choose from the dropdown list.',
        'show_input': True,
        'show_error': True,
        'error_title': 'Invalid Input',
        'error_message': 'Please select a valid value from the dropdown list.',
        'error_type': 'information',
        'ignore_blank': False
    }
    
    assignment_sheet.data_validation(1,0, 200,0, moduleValidation)
    assignment_sheet.data_validation(1,1, 200,1, teacherValidation)

"""
Builds schedule template
"""
def scheduleTemplateBuilder(timetable, timeslots_sheet, writer, lock):
    periods=list(Period.objects.filter(department=timetable.tableYear.department, week=1).order_by('dayNum').values_list('name', flat=True))
    w=timetable.tableYear.department.format.numWeeks
    groups=ModuleGroup.objects.filter(parent__timetable=timetable).order_by('name')
    groupPd=pd.DataFrame(columns=["Name","Period", "Week", "id"])
    if groups:
        for group in groups:
            groupPd.loc[len(groupPd)] = [group.name, group.period.name if group.period else '', group.period.week if group.period else '', group.id]
    groupPd.to_excel(writer, sheet_name="Schedule", index=False)
    
    periodValidation = {
        'validate': 'list',
        'source': periods,
        'input_title': 'Enter a value:',
        'input_message': 'Choose from the dropdown list.',
        'show_input': True,
        'show_error': True,
        'error_title': 'Invalid Input',
        'error_message': 'Please select a valid value from the dropdown list.',
        'error_type': 'information',
        'ignore_blank': True
    }
    weekValidation = {
        'validate': 'integer',
        'criteria': 'between',
        'minimum': 1,
        'maximum': w,
        'input_title': 'Enter a value:',
        'input_message': 'between 1 and ' + str(w),
        'show_input': True,
        'show_error': True,
        'error_title': 'Invalid Input',
        'error_message': 'Please select a valid value from the dropdown list.',
        'error_type': 'information',
        'ignore_blank': True
    }
    timeslots_sheet.set_column('D:D', None, lock, {'hidden': True})
    timeslots_sheet.data_validation(1,1, len(groupPd),1, periodValidation)
    timeslots_sheet.data_validation(1,2, len(groupPd),2, weekValidation)

"""
Read in a workbook and run each reader function
"""
def readInExcel(timetable, ws):
    #Load workbook and get sheet names
    workbook = load_workbook(ws)
    sheets=workbook.sheetnames

    #Define reader functions
    template_readers = {
        'Classes': moduleParentTemplateReader,
        'Teachers': teacherTemplateReader,
        'Schedule': scheduleTemplateReader,
        'Preferences': preferencesTemplateReader,
        'Assignments': assignmentsTemplateReader
    }
    results={}
    #Read each sheet, storing results
    for sheet in sheets:
        results[sheet]=template_readers[sheet](timetable, pd.read_excel(ws, sheet_name=sheet))
        
    return results

"""
reads class template
"""
def moduleParentTemplateReader(timetable, classDF):
    classDF=classDF.replace({0:np.NaN, '':np.NaN, ' ':np.NaN}) #Replace invalid data

    #Iterate through each row of data.
    for i, r in classDF.iterrows():
        #Check all data is valid
        if not pd.isnull(r['Class Name']) and not pd.isnull(r['Num Periods Taught']) and not pd.isnull(r['Num Class Groups']):
            #If object exists try to get the object to update
            if not pd.isnull(r['id']):
                try:
                    cl=ModuleParent.objects.get(id=r['id'])
                except:
                    try:
                        cl=ModuleParent.objects.get(name=r['Class Name'], timetable=timetable)
                    except:
                        continue
            if not cl:
                cl=ModuleParent()
                cl.timetable=timetable
                cl.department=timetable.tableYear.department

            #Save object
            cl.name=r['Class Name']
            cl.numPeriods=r['Num Periods Taught']
            cl.numClasses=r['Num Class Groups']
            cl.save()
    return True

"""
Reads Teacher template
"""
def teacherTemplateReader(timetable, teachersDF):
    teachersDF=teachersDF.replace({0:np.NaN, '':np.NaN, ' ':np.NaN}) #Replace invalid content
    for i, r in teachersDF.iterrows():
        if not pd.isnull(r['Name']) and not pd.isnull(r['Load Per Week']): #Check if data is valid
            if not pd.isnull(r['id']):
                try:
                    teacher=Teacher.objects.get(id=r['id'])
                except:
                    try:
                        teacher=Teacher.objects.get(name=r['Name'], department=timetable.tableYear.department)
                    except:
                        teacher=None
            if not teacher:
                teacher=Teacher()
                teacher.department=timetable.tableYear.department
            teacher.name=r['Name']
            teacher.load=r['Load Per Week']
            teacher.roomNum=r['Room']
            teacher.save()
    return True

"""
Reads schedule template
"""
def scheduleTemplateReader(timetable, timetableDF):
    timetableDF=timetableDF.replace({0:np.NaN, '':np.NaN, ' ':np.NaN}) #Replace invalid
    for i, r in timetableDF.iterrows():
        if not pd.isnull(r['Name']) and not pd.isnull(r['Period']) and not pd.isnull(r['Week']): #Skip invalid data
            try:
                mod=ModuleGroup.objects.get(id=r['id'])
            except:
                try:
                    mod=ModuleGroup.objects.get(name=r['Name'], parent__timetable=timetable)
                except:
                    continue
            
            mod.period=Period.objects.get(department=timetable.tableYear.department, name=r['Period'], week=r['Week'])
            mod.save()
    return True

"""
Reads preference template
"""
def preferencesTemplateReader(timetable, preferencesDF):
    priorities={'NONE':0,'MEDIUM':1,'HIGH':2,'REQUIRED':3}
    preferencesDF=preferencesDF.replace({0:np.NaN, '':np.NaN, ' ':np.NaN}) #Replace invalid
    for i, r in preferencesDF.iterrows():
        if not pd.isnull(r['Class']) and not pd.isnull(r['Teacher']) and not pd.isnull(r['Priority']): #Skip invalid data
            try:
                mods=Module.objects.filter(name=r['Class'], group__parent__timetable=timetable)
                teacher=Teacher.objects.get(name=r['Teacher'], department=timetable.tableYear.department)
            except:
                continue
            try:
                for mod in mods:
                    Preference.objects.create(teacher=teacher, module=mod, priority=priorities[r['Priority']], timetable=timetable)
            except:
                continue       
    return True

"""
Read Assignments template
"""
def assignmentsTemplateReader(timetable, assignmentDF):
    assignmentDF=assignmentDF.replace({0:np.NaN, '':np.NaN, ' ':np.NaN}) #Replace invalid
    for i, r in assignmentDF.iterrows():
        if not pd.isnull(r['Class']) and not pd.isnull(r['Teacher']): #Skip invalid data
            try:
                teacher=Teacher.objects.get(name=r['Teacher'], department=timetable.tableYear.department)
                Module.objects.filter(name=r['Class'], group__parent__timetable=timetable).update(teacher=teacher)                
            except:
                continue
    return True

"""
Builds combing chart export
"""
def combingTemplateBuilder(timetable):
    import os
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    file_path = os.path.join(BASE_DIR, 'static', 'excel_templates', 'combingTemplate.xlsx')

    #Loads pre-defined template
    workbook = load_workbook(filename=file_path)

    #Gets all sheets
    combing_sheet = workbook['Combing']
    teachers_sheet = workbook['Teachers']
    classes_sheet = workbook['Classes']
    format_sheet = workbook['Format']
    assignments_sheet = workbook['Assignments']

    #Appends format data to format sheet
    format_sheet.append([timetable.tableYear.department.format.numPeriods, timetable.tableYear.department.format.numWeeks])

    #Get objects
    teachers=Teacher.objects.filter(department=timetable.tableYear.department)
    classes=ModuleParent.objects.filter(timetable=timetable)
    modules=Module.objects.filter(group__parent__timetable=timetable,group__period__isnull=False, teacher__isnull=False)

    #Append teachers to teacher sheet
    for teacher in teachers:
        teachers_sheet.append([teacher.name, teacher.load])
    
    #Append classes to classes sheet
    for cl in classes:
        classes_sheet.append([cl.name, cl.numPeriods, cl.numClasses])
    
    #Append assignments and set cell color
    for mod in modules:
        assignments_sheet.append([mod.teacher.name, mod.group.period.dayNum, mod.name])
        cell=assignments_sheet.cell(row=assignments_sheet.max_row, column=3)
        cell.fill=PatternFill(fill_type='solid', start_color=mod.group.parent.color.split('#')[1])
    
    #Create an IO object to save workbook and send.
    file = io.BytesIO()
    workbook.save(file)
    file.seek(0)

    return file.getvalue()

"""
Experimental reading from a combing chart function
"""
def readFromCombing(timetable, ws):
    #Read in all sheets to pandas
    combing_sheet = pd.read_excel(ws,sheet_name='Combing')
    teachers_sheet = pd.read_excel(ws,sheet_name='Teachers')
    classes_sheet = pd.read_excel(ws, sheet_name='Classes')
    format_sheet = pd.read_excel(ws, sheet_name='Format')

    #Replace invalid data
    teacherNullCheck=teachers_sheet.replace({"":np.NaN, 0:np.NaN}).copy()
    classesNullCheck=classes_sheet.replace({"":np.NaN, 0:np.NaN}).copy()

    #If any invalid data, return error
    if teacherNullCheck.isnull().sum().sum()>0 or classesNullCheck.isnull().sum().sum()>0:
        return False
    
    #Loop through classes sheet and update
    for i, r in classes_sheet.iterrows():
        try:
            cl=ModuleParent.objects.get(name=r['Class'], timetable=timetable)
        except:
            cl=None
        if not cl:
            cl=ModuleParent()
            cl.timetable=timetable
            cl.department=timetable.tableYear.department
        cl.name=r['Class']
        cl.numPeriods=r['Groups']
        cl.numClasses=r['Periods']
        cl.save()

    #Loop through teachers sheet and update
    for i, r in teachers_sheet.iterrows():
        try:
            teacher=Teacher.objects.get(name=r['Teacher'], department=timetable.tableYear.department)
        except:
            teacher=None
        if not teacher:
            teacher=Teacher()
            teacher.department=timetable.tableYear.department
        teacher.name=r['Teacher']
        teacher.load=r['Load Per Week']
        teacher.save()
    
    #Get format of chart
    periodLen=timetable.tableYear.department.format.numPeriods*5*timetable.tableYear.department.format.numWeeks
    columns=periodLen+2
    readCombing=combing_sheet.replace(0,np.NaN).copy() #Replace any 0's with nan
    readCombing=readCombing.iloc[:,0:columns] #Read in columns
    assignments={i:[] for i in range(1,periodLen+1)}
    sessions={}
    print(assignments)
    count=0
    print(readCombing.head())
    #Try to read in data from chart
    try:
        for index, row in readCombing.iterrows():
            if row['Teacher']=='Class' and row['Load']=="Periods": #Skip teacher and load columns
                break
            else: #Try to get the teacher object of the row
                try:
                    teacher=Teacher.objects.get(name=row['Teacher'], department=timetable.tableYear.department)
                except:
                    continue
            #Iterate through each item in the row. If value exists, add to assignments
            for column_name, value in row.items():
                if column_name not in ['Teacher', 'Load'] and pd.notna(value):
                    assignments[column_name].append((value,teacher))
                    sessions.setdefault(value, 1)

        #Update assignments based on assignments dict.
        for k,v in assignments.items():
            for tup in v:
                Module.objects.filter(group__parent__timetable=timetable, group__session=sessions[tup[0]], name=tup[0]).update(teacher=tup[1])
                sessions[tup[0]]+=1
    except Exception as e:
        print(e)
    return True

"""
Builds calendar export
"""
def exportCalendar(timetable, teacher=None):
    #Get calendar format
    numPeriods=timetable.tableYear.department.format.numPeriods
    weeks=timetable.tableYear.department.format.numWeeks

    #Create IO object and workbook
    calendar = io.BytesIO()
    workbook=xlsxwriter.Workbook(calendar)

    #Create offset for styling. Set a x-coord lookup
    offset=3
    xLookup={'Mon':1, 'Tues':2, 'Wed':3, 'Thurs':4, 'Fri':5}

    #Create formatting
    header_format = workbook.add_format({'bold': True, 'font_size': 14, 'align': 'center', 'valign': 'vcenter'})
    cell_format = workbook.add_format({'font_size': 14, 'text_wrap': True,'align': 'center', 'valign': 'vcenter'})
    title_format = workbook.add_format({'bold': True, 'font_size': 16, 'align': 'center', 'valign': 'vcenter', 'border': 1})
    #Get calendar data for each week. Check if teacher or overall.
    for week in range(1, weeks+1):
        title=timetable.name + " Week " + str(week)
        if teacher:
            classes=Module.objects.filter(group__parent__timetable=timetable, teacher=teacher, group__period__isnull=False, group__period__week=week).order_by('group', 'group__session', 'lesson')
            title=teacher.name + " - " + title
        else:
            classes=ModuleGroup.objects.filter(parent__timetable=timetable, period__isnull=False, period__week=week).order_by('session')
        
        worksheet=workbook.add_worksheet("Week "+str(week)) #Create new sheet for the worksheet
        write_dict={} #To store cell details for writing
   
        #Loop through classes and lookup position
        for cl in classes:
            if teacher:
                coord=cl.group.period.name.split('-')
            else:
                coord=cl.period.name.split('-')
            y=int(coord[1])+offset #get y-coord with offset
            x=xLookup[coord[0]]+offset #Get x-coord with offset

            #Add to dict for each coord
            write_dict.setdefault((y,x),'')
            write_dict[(y,x)]+=(cl.name+'\n\n')

        #Write cell data
        for k,v in write_dict.items():
            worksheet.write(k[0],k[1], v, cell_format)
        
        #Create formula for each period cell
        for i in range(1,numPeriods+1):
            row=i+offset
            worksheet.write_formula(row, offset, str(i), header_format)
        
        #Set headers
        headers=[{'header': 'Period', 'header_format': header_format},
                {'header': 'Mon', 'header_format': header_format},
                {'header': 'Tues', 'header_format': header_format},
                {'header': 'Wed', 'header_format': header_format},
                {'header': 'Thurs','header_format': header_format},
                {'header': 'Fri', 'header_format': header_format},
                ]

        #Add calendar title and format cell heights
        worksheet.merge_range(offset-2, offset, offset-2, offset+5, title, title_format)
        worksheet.set_column(offset+1, offset+5, 30)
        worksheet.set_row(offset-2, 30)

        #Create table
        worksheet.add_table(offset,offset, offset+numPeriods, offset+5, {'columns': headers, 'autofilter':False, 'first_column':True, 'name':'CalWeek'+str(week)})
    
    workbook.close()
    return calendar.getvalue()
    